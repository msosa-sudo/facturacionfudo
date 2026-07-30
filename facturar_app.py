#!/usr/bin/env python3
"""
Facturación Terminales — Anser Indicus SPA
Interfaz web Streamlit v1.0
"""
import io, re, unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuración ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Facturación — Fudo",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Identidad visual Fudo ────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700&display=swap');

/* ── Tipografía — EXCLUYE íconos de Material Symbols ── */
html, body { font-family: 'Barlow', sans-serif !important; }
p, label, input, textarea, select,
[data-testid="stMarkdownContainer"],
[data-testid="stText"],
[data-testid="stAlert"],
[data-testid="stMetricLabel"],
[data-testid="stMetricValue"],
[data-testid="stCaptionContainer"],
button:not([data-testid="stFileUploaderDeleteBtn"]) {
    font-family: 'Barlow', sans-serif !important;
}

/* ── Preservar fuente de íconos Material (evita "uploadupload" y ".arro") ── */
[data-testid="stIconMaterial"],
span[class*="material"],
span[class*="Material"] {
    font-family: 'Material Symbols Rounded', 'Material Icons' !important;
}

/* ── Fondo ── */
[data-testid="stAppViewContainer"] > .main { background-color: #FFFFFF; }

/* ── Títulos ── */
h1 { color: #FF5023 !important; font-weight: 700 !important; }
h2, h3 { color: #3938A0 !important; font-weight: 700 !important; }
h4, h5, h6 { color: #3B3B3B !important; font-weight: 600 !important; }

/* ── Sidebar ── */
[data-testid="stSidebar"] { background-color: #3938A0 !important; }
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #FFFFFF !important; }
[data-testid="stSidebar"] label { color: #FFFFFF !important; font-weight: 600 !important; }
/* Caption / email en sidebar — visible sobre fondo azul */
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
    color: #FFC9BB !important;
    font-size: 12px !important;
}

/* ── Botón primario ── */
button[data-testid="baseButton-primary"] {
    background-color: #FF5023 !important;
    border: none !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    border-radius: 6px !important;
    font-size: 15px !important;
}
button[data-testid="baseButton-primary"]:hover { background-color: #e04010 !important; }

/* ── Botón secundario ── */
button[data-testid="baseButton-secondary"] {
    border: 1.5px solid #3938A0 !important;
    color: #3938A0 !important;
    font-weight: 600 !important;
    border-radius: 6px !important;
}
button[data-testid="baseButton-secondary"]:hover { background-color: #E1E1F5 !important; }

/* ── Botones de descarga ── */
[data-testid="stDownloadButton"] button {
    background-color: #3938A0 !important;
    color: #FFFFFF !important;
    border: none !important;
    font-weight: 700 !important;
    border-radius: 6px !important;
}
[data-testid="stDownloadButton"] button:hover { background-color: #5F5FC4 !important; }
[data-testid="stDownloadButton"] button[data-testid="baseButton-primary"] {
    background-color: #FF5023 !important;
}
[data-testid="stDownloadButton"] button[data-testid="baseButton-primary"]:hover {
    background-color: #e04010 !important;
}

/* ── Métricas ── */
[data-testid="metric-container"] {
    background-color: #E1E1F5 !important;
    border-radius: 8px !important;
    padding: 14px 16px !important;
    border-left: 4px solid #3938A0 !important;
}
[data-testid="stMetricLabel"] label { color: #3938A0 !important; font-weight: 600 !important; }
[data-testid="stMetricValue"]  div  { color: #FF5023 !important; font-weight: 700 !important; }

/* ── Expander: solo borde, sin tocar summary ni íconos ── */
[data-testid="stExpander"] details {
    border: 1.5px solid #FF5023 !important;
    border-radius: 8px !important;
}

/* ── File uploader ── */
[data-testid="stFileUploaderDropzone"] {
    border: 1.5px dashed #5F5FC4 !important;
    border-radius: 6px !important;
    background-color: #FAFAFF !important;
}

/* ── Formulario ── */
[data-testid="stForm"] {
    border: 1.5px solid #E1E1F5 !important;
    border-radius: 10px !important;
}

/* ── Divisor ── */
hr { border-color: #E1E1F5 !important; opacity: 0.6; }

/* ── Caption fuera del sidebar ── */
[data-testid="stCaptionContainer"] p { color: #C6C6C6 !important; }

/* ── Tablas markdown ── */
table { border-collapse: collapse; width: 100%; }
th {
    background-color: #3938A0 !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    padding: 8px 12px !important;
}
td { padding: 6px 12px !important; border-bottom: 1px solid #E1E1F5 !important; }
tr:nth-child(even) td { background-color: #FAFAFF !important; }
</style>
""", unsafe_allow_html=True)

# ─── Acceso por email corporativo ────────────────────────────────────────────
def check_access():
    """
    Pantalla de acceso. Permite entrar a cualquier @fu.do
    o a emails extra listados en st.secrets["allowed_emails"] (separados por coma).
    """
    if st.session_state.get("authenticated"):
        return True

    # Emails adicionales autorizados fuera del dominio @fu.do
    extra_raw   = st.secrets.get("allowed_emails", "")
    extra_emails = {e.strip().lower() for e in extra_raw.split(",") if e.strip()}

    st.markdown("""
    <div style="max-width:440px; margin:80px auto 0;">
        <div style="background:#FF5023; border-radius:10px; padding:28px 32px 18px;
                    margin-bottom:24px; text-align:center;">
            <div style="font-family:'Barlow',sans-serif; font-size:26px;
                        font-weight:700; color:#FFFFFF; letter-spacing:-0.5px;">
                🧾 Facturación Fudo
            </div>
            <div style="font-family:'Barlow',sans-serif; font-size:13px;
                        color:#FFC9BB; margin-top:4px;">
                Anser Indicus SPA — ingresá con tu correo corporativo
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 3, 1])
    with col:
        with st.form("access_form"):
            email = st.text_input("Correo corporativo", placeholder="nombre@fu.do")
            ok    = st.form_submit_button("Ingresar →", use_container_width=True, type="primary")
            if ok:
                email_lower = email.strip().lower()
                if email_lower.endswith("@fu.do") or email_lower in extra_emails:
                    st.session_state["authenticated"] = True
                    st.session_state["user_email"]    = email_lower
                    st.rerun()
                elif not email_lower:
                    st.warning("Ingresá tu correo corporativo.")
                else:
                    st.error("❌ Solo se permite acceso con correos @fu.do")
    return False

# ─── Constantes ───────────────────────────────────────────────────────────────
COMERCIALES   = ['diego toledo', 'jorge de freitas', 'franco de luca', 'milagros sosa vidoni']
PRECIO_UNIT   = 71345
MESES_ES      = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                 7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
PROD_TERM     = '[SPFP] SmartPOS Urovo i9100'
CTA_TERM      = '310114 Venta de Devices Fudo Pagos'
PROD_COM      = 'Comisiones T.O. Plus'
CTA_COM       = '310160 Comisiones Tienda online'

# ─── Helpers ──────────────────────────────────────────────────────────────────
def normalizar(s):
    if not s or str(s).strip() == '' or str(s).lower() == 'nan': return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s)

def es_vacio(val):
    return not val or str(val).strip() == '' or str(val).lower() in ('nan','none','no encontrado','')

def limpiar_rut(rut):
    return str(rut).replace('.','').replace('-','').lower().strip()

def es_comercial(op):
    return any(c in str(op).lower() for c in COMERCIALES)

def parse_fecha_dt(f):
    try: return datetime.strptime(f, '%d/%m/%Y')
    except: return datetime.min

def calc_total_df(df, pu=PRECIO_UNIT):
    if df.empty: return 0
    return sum(round(pu * int(r['cantidad']) * (1 - r['descuento']/100) * 1.19) for _, r in df.iterrows())

# ─── Estilos Excel ───────────────────────────────────────────────────────────
header_fill = PatternFill('solid', start_color='1F4E79')
red_fill    = PatternFill('solid', start_color='FFD7D7')
orange_fill = PatternFill('solid', start_color='FFE0B2')
yellow_fill = PatternFill('solid', start_color='FFFF00')
green_fill  = PatternFill('solid', start_color='E2EFDA')
total_fill  = PatternFill('solid', start_color='D9E1F2')
blue_fill   = PatternFill('solid', start_color='DDEEFF')
field_red   = PatternFill('solid', start_color='FF0000')

def aplicar_header(ws, headers, widths=None):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

# ─── Regiones (bundled en el repo) ───────────────────────────────────────────
@st.cache_data
def cargar_regiones():
    ruta = Path(__file__).parent / "datos" / "Regiones_de_chile_y_comunas.xlsx"
    if not ruta.exists():
        st.warning("⚠️ No se encontró datos/Regiones_de_chile_y_comunas.xlsx — las regiones quedarán vacías.")
        return {}
    df = pd.read_excel(ruta)
    return {normalizar(r['Comuna']): str(r['Validación']).strip()
            for _, r in df.iterrows() if normalizar(str(r['Comuna']))}

def buscar_region(comuna, rl):
    k = normalizar(comuna)
    if not k: return ''
    if k in rl: return rl[k]
    for rk, rv in rl.items():
        if rk and (rk in k or k in rk): return rv
    return ''

def es_comuna_valida(v, rl):
    return buscar_region(v, rl) != ''

# ─── Parsers ──────────────────────────────────────────────────────────────────
def parse_terminal(s):
    s = s.strip().strip('"')
    parts = s.split(':')
    return (parts[1].strip(),
            ':'.join(parts[5:]).strip() if len(parts) > 5 else '',
            [{'qty': int(parts[2].strip()), 'desc': float(parts[3].strip().replace(',','.'))}])

def parse_tfp(s):
    s = s.strip().strip('"')
    parts = s.split(':')
    if len(parts) < 3: return '', '', []
    id_c, nombre = parts[1].strip(), parts[-1].strip()
    ld = {}
    for l in parts[2:]:
        l = l.strip()
        if l == '': break
        if ';' not in l: continue
        try:
            tokens = l.split(';')
            if len(tokens) < 2: continue
            qv = int(tokens[0].strip())
            dv = float(tokens[1].strip().replace(',','.'))
            if qv > 0: ld[dv] = ld.get(dv, 0) + qv
        except (ValueError, IndexError): pass
    return id_c, nombre, [{'qty': q, 'desc': d} for d, q in sorted(ld.items())]

def get_ref_string(row, col_desc, col_extref, tiene_extref):
    if tiene_extref:
        v = str(row.get(col_extref, '') or '').strip()
        if v.startswith(('Terminal:', 'TFP:')): return v
    v = str(row.get(col_desc, '') or '').strip().strip('"')
    if v.startswith(('Terminal:', 'TFP:')): return v
    return None

# ─── Cargadores de datos ──────────────────────────────────────────────────────
def leer_collection(f):
    try:
        if f.name.lower().endswith('.csv'):
            content = f.read().decode('utf-8', errors='replace')
            f.seek(0)
            sep = ';' if content.count(';') > content.count(',') else ','
            df_c = pd.read_csv(io.StringIO(content), sep=sep, header=None, on_bad_lines='skip', engine='python')
        else:
            df_c = pd.read_excel(f, header=None, engine='calamine')
    except Exception:
        f.seek(0)
        df_c = pd.read_excel(f, header=None, engine='openpyxl')
    df_c.columns = df_c.iloc[0]
    df_c = df_c[1:].reset_index(drop=True)
    col_desc   = 'Descripción de la operación (reason)'
    col_opid   = 'Número de operación de Mercado Pago (operation_id)'
    col_monto  = 'Valor del producto (transaction_amount)'
    col_fecha  = 'Fecha de compra (date_created)'
    col_op     = 'Operador en cobros de Point (operator_name)'
    col_extref = 'Código de referencia (external_reference)'
    missing = [c for c in [col_desc, col_opid, col_monto] if c not in df_c.columns]
    if missing:
        raise ValueError(f"Columnas faltantes en collection: {', '.join(missing)}")
    df_c[col_monto] = pd.to_numeric(df_c[col_monto].astype(str).str.replace(',','.'), errors='coerce')
    df_c[col_opid]  = df_c[col_opid].astype(str).str.strip()
    cols = {'desc': col_desc, 'opid': col_opid, 'monto': col_monto,
            'fecha': col_fecha, 'op': col_op, 'extref': col_extref}
    return df_c, cols

def _cols_disponibles(df, n=8):
    return ', '.join(f'"{c}"' for c in list(df.columns)[:n])

def leer_billing(f):
    df = pd.read_csv(f, dtype=str)
    # Normalizar nombre de columnas (strip espacios)
    df.columns = df.columns.str.strip()
    requeridas = ['ID', 'Código', 'Razón social', 'Nombre', 'Giro', 'Domicilio', 'Comuna', 'Email']
    faltantes  = [c for c in requeridas if c not in df.columns]
    if faltantes:
        raise ValueError(
            f"**Billing data**: columna(s) no encontrada(s): {', '.join(faltantes)}\n\n"
            f"Columnas que tiene el archivo: {_cols_disponibles(df)}\n\n"
            f"Verificá que estés subiendo el archivo correcto (billing_data_FECHA.csv de dash.fu.do)"
        )
    df['RUT_clean'] = (df['Código'].astype(str)
                       .str.replace('.', '', regex=False)
                       .str.replace('-', '', regex=False)
                       .str.strip().str.lower())
    df['ID'] = df['ID'].astype(str).str.strip()
    return df.set_index('ID')[['RUT_clean','Razón social','Nombre','Giro','Domicilio','Comuna','Email']].to_dict('index')

def leer_contactos(f):
    df = pd.read_excel(f, dtype=str)
    df.columns = df.columns.str.strip()
    requeridas = ['ID', 'Referencia', 'NIF', 'Nombre']
    faltantes  = [c for c in requeridas if c not in df.columns]
    if faltantes:
        raise ValueError(
            f"**Contactos Odoo**: columna(s) no encontrada(s): {', '.join(faltantes)}\n\n"
            f"Columnas que tiene el archivo: {_cols_disponibles(df)}\n\n"
            f"Verificá que sea la exportación de res.partner de Odoo"
        )
    df['Referencia_clean'] = df['Referencia'].astype(str).str.strip()
    df['NIF_clean'] = (df['NIF'].astype(str)
                       .str.replace('.', '', regex=False)
                       .str.replace('-', '', regex=False)
                       .str.strip().str.lower())
    df['DB_ID'] = df['ID'].astype(str).str.extract(r'res_partner_(\d+)_')
    _cols_odoo = {'razon_social': 'Nombre', 'giro': 'Giro',
                  'domicilio': 'Nombre de la calle', 'comuna': 'Ciudad'}
    ref_to_odoo_datos = {}
    for _, r in df.iterrows():
        ref = str(r.get('Referencia', '')).strip()
        if ref and ref != 'nan':
            ref_to_odoo_datos[ref] = {
                c: str(r.get(co, '') or '').strip()
                for c, co in _cols_odoo.items() if co in df.columns
            }
    return {
        'ref_to_dbid':       df.set_index('Referencia_clean')['DB_ID'].to_dict(),
        'ref_to_nif':        df.set_index('Referencia_clean')['NIF_clean'].to_dict(),
        'ref_to_nombre':     df.set_index('Referencia_clean')['Nombre'].to_dict(),
        'nif_to_dbid':       df.set_index('NIF_clean')['DB_ID'].to_dict(),
        'nif_to_ref':        df.set_index('NIF_clean')['Referencia_clean'].to_dict(),
        'ref_to_odoo_datos': ref_to_odoo_datos,
    }

def leer_odoo(f):
    df = pd.read_excel(f)
    df.columns = df.columns.str.strip()
    posibles = ['Referencia del pago', 'Referencia de pago', 'Referencia']
    rc = next((c for c in posibles if c in df.columns), None)
    if rc is None:
        raise ValueError(
            f"**Asiento contable**: no se encontró columna de referencia de pago.\n\n"
            f"Columnas disponibles: {_cols_disponibles(df)}\n\n"
            f"Verificá que sea la exportación de asientos contables de Odoo"
        )
    return set(df[rc].astype(str).str.replace("'", "").str.strip().tolist())

def leer_accounts(f):
    df = pd.read_csv(f, skiprows=3, encoding='utf-8-sig',
                     on_bad_lines='skip', engine='python', quotechar='"')
    sa, sn = {}, {}
    for _, r in df.iterrows():
        n = str(r.get('Nombre', '') or '').strip()
        if n and n.lower() != 'nan':
            s = re.sub(r'\s+', '', n.lower())
            sa[s] = str(r['ID']).strip()
            sn[s] = n
    return sa, sn

# ─── get_billing ──────────────────────────────────────────────────────────────
def get_billing(id_cuenta, billing_raw, rl):
    datos = billing_raw.get(str(id_cuenta).strip(), {})
    if not datos: return datos
    comuna    = str(datos.get('Comuna', '') or '')
    domicilio = str(datos.get('Domicilio', '') or '')
    if not es_comuna_valida(comuna, rl) and es_comuna_valida(domicilio, rl):
        datos = dict(datos)
        datos['Comuna'], datos['Domicilio'] = domicilio, comuna
    return datos

# ─── Procesamiento principal ──────────────────────────────────────────────────
def procesar(df_c, cols, billing_raw, refs, ids_facturados, rl,
             slug_to_accid, slug_to_nombre, hacer_terminales, hacer_comisiones):
    col_desc   = cols['desc']
    col_opid   = cols['opid']
    col_monto  = cols['monto']
    col_fecha  = cols['fecha']
    col_op     = cols['op']
    col_extref = cols['extref']
    tiene_extref = col_extref in df_c.columns

    ref_to_dbid  = refs['ref_to_dbid']
    ref_to_nif   = refs['ref_to_nif']
    nif_to_dbid  = refs['nif_to_dbid']
    ref_to_nombre = refs['ref_to_nombre']

    # Máscaras
    mask_reason  = df_c[col_desc].astype(str).str.startswith(('Terminal:', 'TFP:'))
    mask_extref  = (df_c[col_extref].astype(str).str.startswith(('Terminal:', 'TFP:'))
                    if tiene_extref else pd.Series([False]*len(df_c), index=df_c.index))
    mask_op      = (df_c[col_op].apply(es_comercial) if col_op in df_c.columns
                    else pd.Series([False]*len(df_c), index=df_c.index))
    mask_comision = df_c[col_desc].astype(str).str.contains('Deuda por comisiones', na=False)
    df_term = df_c[mask_reason | mask_extref | mask_op | mask_comision].copy().reset_index(drop=True)

    rows, rows_comision = [], []
    duplicados, alertas_monto, alertas_operador, alertas_formato = [], [], [], []
    comisiones_sin_cuenta = []

    for _, row in df_term.iterrows():
        desc_val   = str(row[col_desc]).strip().strip('"')
        opid       = row[col_opid]
        monto_real = row[col_monto]
        operador   = str(row.get(col_op, '')) if col_op in df_c.columns else ''

        if opid in ids_facturados:
            duplicados.append(desc_val)
            continue

        # ── Comisiones ────────────────────────────────────────────
        if 'Deuda por comisiones' in desc_val:
            if not hacer_comisiones:
                continue
            fecha_c  = str(row.get(col_fecha, '')).strip().split(' ')[0]
            if fecha_c == 'nan': fecha_c = ''
            extref_v = str(row.get(col_extref, '') or '').strip() if tiene_extref else ''
            slug = extref_v.split('@', 1)[1].lower() if '@' in extref_v else ''
            acc_id = slug_to_accid.get(slug, '') if slug else ''
            if not acc_id:
                comisiones_sin_cuenta.append({
                    'extref': extref_v, 'slug': slug,
                    'operation_id': opid, 'monto': monto_real, 'fecha': fecha_c
                })
                continue
            billing_c = get_billing(acc_id, billing_raw, rl)
            rut_c     = billing_c.get('RUT_clean', 'NO ENCONTRADO')
            sin_dat_c = (rut_c == 'NO ENCONTRADO')
            db_id_c   = ref_to_dbid.get(acc_id, '')
            if not db_id_c:
                rk = limpiar_rut(rut_c) if rut_c != 'NO ENCONTRADO' else ''
                db_id_c = nif_to_dbid.get(rk, 'ND') if rk else 'ND'
            try:
                _dt = datetime.strptime(fecha_c, '%d/%m/%Y')
                terminos = f"'{MESES_ES[_dt.month]} {_dt.year} -  DF"
            except Exception:
                terminos = ''
            rows_comision.append({
                'id_cuenta': acc_id,
                'nombre_cuenta': slug_to_nombre.get(slug, acc_id),
                'operation_id': opid, 'monto_real': monto_real,
                'precio_sin_iva': monto_real / 1.19,
                'RUT_billing': rut_c, 'RUT_odoo': ref_to_nif.get(acc_id, ''),
                'razon_social': billing_c.get('Razón social', ''),
                'nombre_billing': billing_c.get('Nombre', ''),
                'giro': billing_c.get('Giro', ''),
                'domicilio': billing_c.get('Domicilio', ''),
                'comuna': billing_c.get('Comuna', ''),
                'email': billing_c.get('Email', ''),
                'db_id': db_id_c, 'sin_datos': sin_dat_c,
                'es_consumidor_final': limpiar_rut(rut_c) == '111111111',
                'fecha_compra': fecha_c, 'terminos': terminos,
            })
            continue

        # ── Terminales ────────────────────────────────────────────
        if not hacer_terminales:
            continue
        ref_str   = get_ref_string(row, col_desc, col_extref, tiene_extref)
        tiene_ref = ref_str is not None

        if not tiene_ref and es_comercial(operador):
            alertas_operador.append({
                'operador': operador.strip(), 'descripcion': desc_val,
                'operation_id': opid, 'monto': monto_real,
                'fecha': str(row.get(col_fecha, '')).split(' ')[0]
            })
            continue

        if not tiene_ref:
            alertas_formato.append({
                'descripcion': desc_val,
                'extref': str(row.get(col_extref, '')) if tiene_extref else '',
                'operation_id': opid, 'monto': monto_real,
                'fecha': str(row.get(col_fecha, '')).split(' ')[0]
            })
            continue

        if ref_str.startswith('Terminal:'):
            id_c, nombre, lineas = parse_terminal(ref_str)
        elif ref_str.startswith('TFP:'):
            id_c, nombre, lineas = parse_tfp(ref_str)
            if not lineas: continue
        else:
            continue

        monto_esp       = round(sum(PRECIO_UNIT * l['qty'] * (1 - l['desc']/100) * 1.19 for l in lineas))
        monto_diferente = abs(monto_real - monto_esp) > 1
        if monto_diferente:
            alertas_monto.append({
                'nombre': nombre, 'id_cuenta': id_c, 'operation_id': opid,
                'monto_esperado': monto_esp, 'monto_real': monto_real,
                'diferencia': monto_real - monto_esp
            })

        billing     = get_billing(id_c, billing_raw, rl)
        rut_billing = billing.get('RUT_clean', 'NO ENCONTRADO')
        rut_odoo    = ref_to_nif.get(id_c, '')
        db_id       = ref_to_dbid.get(id_c, '')
        if not db_id:
            rk = limpiar_rut(rut_billing) if rut_billing != 'NO ENCONTRADO' else ''
            db_id = nif_to_dbid.get(rk, 'ND') if rk else 'ND'
        razon       = billing.get('Razón social', '')
        sin_datos   = (rut_billing == 'NO ENCONTRADO')
        es_cf       = limpiar_rut(rut_billing) == '111111111'
        fecha_t     = str(row.get(col_fecha, '')).strip().split(' ')[0]
        if fecha_t == 'nan': fecha_t = ''

        for linea in lineas:
            rows.append({
                'id_cuenta': id_c, 'cantidad': linea['qty'], 'descuento': linea['desc'],
                'nombre_cuenta': nombre, 'operation_id': opid, 'monto': monto_real,
                'RUT_billing': rut_billing, 'RUT_odoo': rut_odoo,
                'razon_social': razon, 'nombre_billing': billing.get('Nombre', ''),
                'giro': billing.get('Giro', ''), 'domicilio': billing.get('Domicilio', ''),
                'comuna': billing.get('Comuna', ''), 'email': billing.get('Email', ''),
                'db_id': db_id, 'contacto_nombre': ref_to_nombre.get(id_c, ''),
                'monto_diferente': monto_diferente, 'sin_datos': sin_datos,
                'es_consumidor_final': es_cf, 'fecha_compra': fecha_t,
            })

    return {
        'rows': rows, 'rows_comision': rows_comision,
        'duplicados': duplicados, 'alertas_monto': alertas_monto,
        'alertas_operador': alertas_operador, 'alertas_formato': alertas_formato,
        'comisiones_sin_cuenta': comisiones_sin_cuenta,
    }

# ─── Clasificación de contactos ───────────────────────────────────────────────
def clasificar_contactos(df_work, df_comision, refs, rl):
    ref_to_dbid      = refs['ref_to_dbid']
    ref_to_nif       = refs['ref_to_nif']
    nif_to_dbid      = refs['nif_to_dbid']
    nif_to_ref       = refs['nif_to_ref']
    ref_to_odoo_datos = refs['ref_to_odoo_datos']

    frames = []
    if not df_work.empty:    frames.append(df_work[~df_work['sin_datos']])
    if not df_comision.empty: frames.append(df_comision[~df_comision['sin_datos']])
    if not frames:
        return [], [], [], [], []
    df_all = pd.concat(frames, ignore_index=True).drop_duplicates('id_cuenta')

    casos_ok, casos_dc, casos_act, casos_crear, casos_rut_otro = [], [], [], [], []

    for _, row in df_all.iterrows():
        id_c        = row['id_cuenta']
        rut_billing = row['RUT_billing']
        rut_odoo    = row['RUT_odoo']
        ref_en_odoo = id_c in ref_to_dbid
        rut_en_odoo = limpiar_rut(rut_billing) in nif_to_dbid

        if ref_en_odoo:
            if limpiar_rut(rut_billing) == limpiar_rut(rut_odoo):
                casos_ok.append(id_c)
                datos_odoo   = ref_to_odoo_datos.get(id_c, {})
                campos_b     = {'razon_social': row['razon_social'], 'giro': row['giro'],
                                'domicilio': row['domicilio'], 'comuna': row['comuna']}
                diffs = {}
                for campo, val_b in campos_b.items():
                    val_o = datos_odoo.get(campo, '')
                    if not es_vacio(val_b) and normalizar(val_b) != normalizar(val_o):
                        diffs[campo] = {'billing': val_b, 'odoo': val_o}
                if diffs:
                    casos_dc.append({
                        'id_cuenta': id_c, 'nombre_cuenta': row['nombre_cuenta'],
                        'db_id': row['db_id'], 'RUT': rut_billing,
                        'razon_social': row['razon_social'], 'diffs': diffs
                    })
            else:
                casos_act.append({
                    'id_cuenta': id_c, 'nombre_cuenta': row['nombre_cuenta'],
                    'RUT_billing': rut_billing, 'RUT_odoo': rut_odoo,
                    'razon_social': row['razon_social'], 'db_id': row['db_id'],
                    'giro': row['giro'], 'domicilio': row['domicilio'],
                    'comuna': row['comuna'], 'email': row['email'],
                    'region': buscar_region(row['comuna'], rl)
                })
        elif rut_en_odoo:
            rut_clean  = limpiar_rut(rut_billing)
            ref_exist  = nif_to_ref.get(rut_clean, '')
            dbid_exist = nif_to_dbid.get(rut_clean, '')
            casos_rut_otro.append({
                'id_cuenta': id_c, 'nombre_cuenta': row['nombre_cuenta'],
                'RUT': rut_billing, 'razon_social': row['razon_social'],
                'db_id_existente': dbid_exist, 'ref_existente': ref_exist
            })
            casos_crear.append({
                'id_cuenta': id_c, 'nombre_cuenta': row['nombre_cuenta'],
                'RUT': rut_billing, 'razon_social': row['razon_social'],
                'giro': row['giro'], 'domicilio': row['domicilio'],
                'comuna': row['comuna'], 'email': row['email'],
                'region': buscar_region(row['comuna'], rl),
                'rut_ya_existe': True,
                'ref_existente': ref_exist, 'db_id_existente': dbid_exist
            })
        else:
            casos_crear.append({
                'id_cuenta': id_c, 'nombre_cuenta': row['nombre_cuenta'],
                'RUT': rut_billing, 'razon_social': row['razon_social'],
                'giro': row['giro'], 'domicilio': row['domicilio'],
                'comuna': row['comuna'], 'email': row['email'],
                'region': buscar_region(row['comuna'], rl)
            })

    return casos_ok, casos_dc, casos_act, casos_crear, casos_rut_otro

# ─── Generación Excel de Contactos ───────────────────────────────────────────
def generar_excel_contactos(casos_crear, casos_act, casos_rut_otro, casos_dc, rl):
    wb = Workbook()
    cont_headers  = ['Nombre','Tipo de compañía','Empresa relacionada','Nombre de la calle','Ciudad',
                     'Provincia','Idioma','País','Tipo de identificación','NIF','Tipo de contribuyente',
                     'Giro','Correo DTE','Correo electrónico','Referencia','Compañía','Enlace a página web']
    campos_criticos = {1:'razon_social',4:'domicilio',5:'comuna',6:'region',10:'RUT',12:'giro',13:'email',14:'email'}

    ws = wb.active; ws.title = 'Importación Clientes CL'
    aplicar_header(ws, cont_headers, [40,15,20,35,20,25,18,10,20,15,25,30,35,35,15,20,40])
    h18 = ws.cell(row=1, column=18, value='⚠ Nota (no importar)')
    h18.fill = PatternFill('solid', start_color='FFF2CC')
    h18.font = Font(bold=True, name='Arial', size=10, color='7F6000')
    ws.column_dimensions['R'].width = 55
    h19 = ws.cell(row=1, column=19, value='Validación Comuna SII')
    h19.fill = PatternFill('solid', start_color='D9E1F2')
    h19.font = Font(bold=True, name='Arial', size=10)
    h19.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['S'].width = 22

    # Dedup por id_cuenta
    id_vistos, casos_dedup = set(), []
    for c in casos_crear:
        if c['id_cuenta'] not in id_vistos:
            id_vistos.add(c['id_cuenta']); casos_dedup.append(c)

    for row_idx, c in enumerate(casos_dedup, 2):
        rut_d    = c['RUT'] if c['RUT'] != 'NO ENCONTRADO' else ''
        es_caso4 = c.get('rut_ya_existe', False)
        data = [c['razon_social'],'Compañía','',c['domicilio'],c['comuna'],c['region'],
                'Spanish / Español','Chile','RUT',rut_d,'IVA afecto 1ª categoría',
                c['giro'],c['email'],c['email'],c['id_cuenta'],'Anser Indicus SPA',
                f"https://dash.fu.do/accounts/{c['id_cuenta']}"]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(vertical='center')
            if es_caso4: cell.fill = yellow_fill
        if es_caso4:
            nota = (f"⚠ Ya existe otra Referencia en Odoo "
                    f"(Ref: {c.get('ref_existente','?')} | DB_ID: {c.get('db_id_existente','?')}). "
                    f"Verificar si es el mismo cliente.")
            cn = ws.cell(row=row_idx, column=18, value=nota)
            cn.font = Font(name='Arial', size=9, color='7F6000')
            cn.fill = yellow_fill
            cn.alignment = Alignment(vertical='center', wrap_text=True)
        # Validación SII
        com_v = str(c['comuna'] or '').strip()
        if len(com_v) > 20:
            cv = ws.cell(row=row_idx, column=19, value='SUPERA LÍMITE SII')
            cv.fill = PatternFill('solid', start_color='FF0000')
            cv.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
            cv.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=5).fill = field_red
            ws.cell(row=row_idx, column=5).font = Font(name='Arial', size=10, color='FFFFFF', bold=True)
        else:
            cv = ws.cell(row=row_idx, column=19, value='OK')
            cv.fill = green_fill
            cv.font = Font(name='Arial', size=9, color='375623')
            cv.alignment = Alignment(horizontal='center', vertical='center')
        # Campos críticos vacíos
        row_data = {'razon_social':c['razon_social'],'domicilio':c['domicilio'],
                    'comuna':c['comuna'],'region':c['region'],'RUT':c['RUT'],
                    'giro':c['giro'],'email':c['email']}
        for col_idx, campo in campos_criticos.items():
            if es_vacio(row_data.get(campo, '')):
                ws.cell(row=row_idx, column=col_idx).fill = field_red
                ws.cell(row=row_idx, column=col_idx).font = Font(name='Arial', size=10, color='FFFFFF', bold=True)

    ws.cell(row=len(casos_dedup)+3, column=1, value='NOTAS:').font = Font(bold=True, name='Arial')
    ws.cell(row=len(casos_dedup)+3, column=2,
        value='🔴 Fila roja = sin RUT | 🟥 Celda roja = campo vacío | 🟨 Fila amarilla = RUT ya existe | 🔴S = SUPERA LÍMITE SII'
    ).font = Font(name='Arial', color='CC5500')

    # Contactos a actualizar + sección en importación
    if casos_act:
        ws_act = wb.create_sheet('⚠ Contactos a actualizar')
        aplicar_header(ws_act,
            ['ID Cuenta','Nombre en MP','Razón Social (billing)','RUT billing','RUT en Odoo','DB_ID Odoo','Acción requerida'],
            [12,30,35,18,18,12,55])
        ws_act.cell(row=1, column=7).fill = PatternFill('solid', start_color='CC0000')
        for row_idx, c in enumerate(casos_act, 2):
            accion = f"1) Cambiar Referencia a '{c['id_cuenta']}_old'  |  2) Crear contacto nuevo"
            for col_idx, val in enumerate([c['id_cuenta'],c['nombre_cuenta'],c['razon_social'],
                                           c['RUT_billing'],c['RUT_odoo'],c['db_id'],accion], 1):
                cell = ws_act.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.fill = red_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            ws_act.row_dimensions[row_idx].height = 35
        # Agregar sección en hoja principal
        next_row = len(casos_dedup) + 5
        ws.cell(row=next_row, column=1,
            value='CONTACTOS A ACTUALIZAR:').font = Font(bold=True, name='Arial', size=10, color='CC0000')
        next_row += 1
        for c in casos_act:
            rut_d = c['RUT_billing'] if c['RUT_billing'] != 'NO ENCONTRADO' else ''
            data = [c['razon_social'],'Compañía','',c['domicilio'],c['comuna'],c['region'],
                    'Spanish / Español','Chile','RUT',rut_d,'IVA afecto 1ª categoría',
                    c['giro'],c['email'],c['email'],c['id_cuenta'],'Anser Indicus SPA',
                    f"https://dash.fu.do/accounts/{c['id_cuenta']}"]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.fill = orange_fill
                cell.alignment = Alignment(vertical='center')
            com_act = str(c['comuna'] or '').strip()
            if len(com_act) > 20:
                cv2 = ws.cell(row=next_row, column=19, value='SUPERA LÍMITE SII')
                cv2.fill = PatternFill('solid', start_color='FF0000')
                cv2.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
                cv2.alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=next_row, column=5).fill = field_red
            else:
                cv2 = ws.cell(row=next_row, column=19, value='OK')
                cv2.fill = green_fill
                cv2.font = Font(name='Arial', size=9, color='375623')
                cv2.alignment = Alignment(horizontal='center', vertical='center')
            next_row += 1

    # RUT existe otro ID
    if casos_rut_otro:
        ws_rut = wb.create_sheet('⚠ RUT existe - otro ID')
        aplicar_header(ws_rut,
            ['ID Cuenta (nuevo)','Nombre en MP','RUT','Razón Social','DB_ID existente','Referencia existente','Acción'],
            [14,30,18,35,14,20,45])
        for row_idx, c in enumerate(casos_rut_otro, 2):
            accion = f"RUT ya existe con Referencia '{c['ref_existente']}'. Verificar si es el mismo cliente."
            for col_idx, val in enumerate([c['id_cuenta'],c['nombre_cuenta'],c['RUT'],
                                           c['razon_social'],c['db_id_existente'],c['ref_existente'],accion], 1):
                cell = ws_rut.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.fill = blue_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            ws_rut.row_dimensions[row_idx].height = 35

    # Mismo RUT varios IDs
    rut_counts = {}
    for c in casos_crear:
        rk = limpiar_rut(c['RUT'])
        rut_counts.setdefault(rk, []).append(c)
    rut_multi = {k: v for k, v in rut_counts.items() if len(v) > 1}
    if rut_multi:
        ws_m = wb.create_sheet('⚠ Mismo RUT - varios IDs')
        aplicar_header(ws_m, ['RUT','Razón Social','IDs de cuenta','Nombres en MP'], [18,35,30,40])
        for row_idx, (rk, items) in enumerate(rut_multi.items(), 2):
            ids    = [i['id_cuenta'] for i in items]
            noms   = [i['nombre_cuenta'] for i in items]
            for col_idx, val in enumerate([rk, items[0]['razon_social'], ', '.join(ids), ', '.join(noms)], 1):
                cell = ws_m.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.fill = orange_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            ws_m.row_dimensions[row_idx].height = 30

    # Datos cambiados
    if casos_dc:
        ws_dc = wb.create_sheet('⚠ Datos actualizados')
        aplicar_header(ws_dc,
            ['ID Cuenta','DB_ID Odoo','RUT','Nombre','Campo','Valor en billing','Valor en Odoo'],
            [14,12,18,30,18,40,40])
        ws_dc.cell(row=1, column=1).fill = PatternFill('solid', start_color='FF6600')
        fila_dc = 2
        labels = {'razon_social':'Razón Social','giro':'Giro','domicilio':'Domicilio','comuna':'Comuna'}
        for c in casos_dc:
            primera = True
            for campo, vals in c['diffs'].items():
                rv = [c['id_cuenta'] if primera else '', c['db_id'] if primera else '',
                      c['RUT'] if primera else '', c['nombre_cuenta'] if primera else '',
                      labels.get(campo, campo), vals['billing'], vals['odoo']]
                for col_idx, val in enumerate(rv, 1):
                    cell = ws_dc.cell(row=fila_dc, column=col_idx, value=val)
                    cell.font = Font(name='Arial', size=10); cell.fill = yellow_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                fila_dc += 1; primera = False

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output

# ─── Generación Excel de Facturación ─────────────────────────────────────────
def generar_excel_facturacion(df_work, rows_comision, alertas_monto, alertas_operador, alertas_formato, comisiones_sin_cuenta=None):
    fecha_hoy = date.today().strftime('%d/%m/%Y')
    wb = Workbook()

    # ── Hoja Terminales ──
    fact_headers = ['Orden','Contacto/Id. de la DB','Referencia','Fecha de Factura/Recibo','Fecha vencimiento',
                    'Referencia de Pago','Diario','Tipo de Documento','Líneas de factura/Producto',
                    'Líneas de factura/Cuenta','Líneas de factura/Cantidad mínima',
                    'Líneas de factura/Precio unitario','Líneas de factura/Impuesto',
                    'Líneas de factura/Descuento (%)']
    ws1 = wb.active; ws1.title = 'Terminales'
    aplicar_header(ws1, fact_headers, [8,15,35,22,18,25,20,20,30,38,12,15,30,12])

    orden = 1; prev_opid = None
    for row_idx, (_, row) in enumerate(df_work.iterrows(), 2):
        es_primera = row['operation_id'] != prev_opid
        if es_primera: orden_val = orden; orden += 1; prev_opid = row['operation_id']
        else: orden_val = ''
        tipo_doc    = 'Boleta Electrónica' if row['es_consumidor_final'] else 'Factura Electrónica'
        nombre_ref  = row['nombre_billing'] if row['nombre_billing'] else row['nombre_cuenta']
        contacto_v  = row['db_id'] if row['db_id'] != 'ND' else ''
        if es_primera:
            data = [orden_val, contacto_v, f"Terminales - {nombre_ref}",
                    fecha_hoy, fecha_hoy, f"'{row['operation_id']}",
                    'Factura Electrónica', tipo_doc, PROD_TERM, CTA_TERM,
                    int(row['cantidad']), PRECIO_UNIT, 'IVA 19 Venta', row['descuento']]
        else:
            data = ['','','','','','','','', PROD_TERM, CTA_TERM,
                    int(row['cantidad']), PRECIO_UNIT, 'IVA 19 Venta', row['descuento']]
        for col_idx, val in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Arial', size=10); cell.alignment = Alignment(vertical='center')
        if row['sin_datos']:
            for c in range(1, 15): ws1.cell(row=row_idx, column=c).fill = red_fill
        elif row['db_id'] == 'ND':
            ws1.cell(row=row_idx, column=2).fill = orange_fill
        if row['monto_diferente']:
            ws1.cell(row=row_idx, column=6).fill = yellow_fill
            ws1.cell(row=row_idx, column=11).fill = yellow_fill

    # ── Hoja Comisiones ──
    if rows_comision:
        ws_com = wb.create_sheet('Comisiones')
        com_headers = ['Orden','Contacto/Id. de la DB','Referencia',
                       'Fecha de Factura/Recibo','Fecha vencimiento','Referencia de Pago',
                       'Términos y condiciones','Diario','Tipo de Documento',
                       'Líneas de factura/Producto','Líneas de factura/Cuenta',
                       'Líneas de factura/Cantidad mínima','Líneas de factura/Precio unitario',
                       'Líneas de factura/Impuesto','Líneas de factura/Descuento (%)']
        aplicar_header(ws_com, com_headers, [8,15,32,22,18,25,25,20,20,22,35,12,20,15,12])
        for row_idx, rc in enumerate(rows_comision, 2):
            tipo_doc_c   = 'Boleta Electrónica' if rc['es_consumidor_final'] else 'Factura Electrónica'
            nombre_ref_c = rc['nombre_billing'] if rc['nombre_billing'] else rc['nombre_cuenta']
            data_c = ['', rc['db_id'] if rc['db_id'] != 'ND' else '',
                      nombre_ref_c, fecha_hoy, fecha_hoy, f"'{rc['operation_id']}",
                      rc['terminos'], 'Factura Electrónica', tipo_doc_c,
                      PROD_COM, CTA_COM, 1, rc['precio_sin_iva'], 'IVA 19 Venta', 0]
            for col_idx, val in enumerate(data_c, 1):
                cell = ws_com.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.alignment = Alignment(vertical='center')
            if rc['sin_datos']:
                for c in range(1, 16): ws_com.cell(row=row_idx, column=c).fill = red_fill
            elif rc['db_id'] == 'ND':
                ws_com.cell(row=row_idx, column=2).fill = orange_fill

    # ── Hoja Comisiones sin cuenta ──
    csc = comisiones_sin_cuenta or []
    if csc:
            ws_ce = wb.create_sheet('⚠ Comisiones sin cuenta')
            aplicar_header(ws_ce,
                ['External Reference','Slug extraído','Operation ID','Monto','Fecha','Acción'],
                [35,25,20,14,14,45])
            for row_idx, e in enumerate(csc, 2):
                for col_idx, val in enumerate([e['extref'],e['slug'],f"'{e['operation_id']}",
                                               e['monto'],e['fecha'],
                                               "No se encontró cuenta en accounts.csv — verificar manual"], 1):
                    cell = ws_ce.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = Font(name='Arial', size=10); cell.fill = red_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                ws_ce.row_dimensions[row_idx].height = 30

    # ── Hoja Resumen ──
    ws_r = wb.create_sheet('Resumen')
    ws_r.column_dimensions['A'].width = 38
    ws_r.column_dimensions['B'].width = 48
    ws_r.column_dimensions['C'].width = 18
    ws_r.column_dimensions['D'].width = 28
    t1 = ws_r.cell(row=1, column=1, value='RESUMEN DE FACTURACIÓN')
    t1.font = Font(bold=True, name='Arial', size=13, color='FFFFFF')
    t1.fill = header_fill; t1.alignment = Alignment(horizontal='center', vertical='center')
    ws_r.merge_cells('A1:C1'); ws_r.row_dimensions[1].height = 24
    ws_r.cell(row=2, column=1, value=f"Fecha: {fecha_hoy}").font = Font(italic=True, name='Arial', size=10, color='666666')

    def bloque(fila, label, valor, fill, nota=''):
        c1 = ws_r.cell(row=fila, column=1, value=label)
        c1.font = Font(bold=True, name='Arial', size=11); c1.fill = fill
        c1.alignment = Alignment(vertical='center', horizontal='right')
        c1.border = Border(outline=Side(style='thin'))
        c2 = ws_r.cell(row=fila, column=2, value=valor)
        c2.font = Font(bold=True, name='Arial', size=12, color='1F4E79'); c2.fill = fill
        c2.number_format = '$#,##0'; c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border = Border(outline=Side(style='thin'))
        if nota: ws_r.cell(row=fila, column=3, value=nota).font = Font(italic=True, name='Arial', size=9, color='444444')

    total_term = calc_total_df(df_work)
    df_ok_t    = df_work[~df_work['sin_datos']] if not df_work.empty else pd.DataFrame()
    df_nd_t    = df_work[df_work['sin_datos']]  if not df_work.empty else pd.DataFrame()
    total_com  = round(sum(r['monto_real'] for r in rows_comision)) if rows_comision else 0
    bloque(4, 'TOTAL COMPLETO (con IVA)', total_term + total_com, total_fill)
    if not df_work.empty:
        bloque(6, 'Terminales — listas para importar', calc_total_df(df_ok_t), green_fill,
               f'→ {df_ok_t["operation_id"].nunique()} facturas' if not df_ok_t.empty else '')
        bloque(7, 'Terminales — sin datos', calc_total_df(df_nd_t), red_fill,
               f'→ {df_nd_t["operation_id"].nunique()} facturas' if not df_nd_t.empty else '')
    if rows_comision:
        com_ok = [r for r in rows_comision if not r['sin_datos']]
        com_nd = [r for r in rows_comision if r['sin_datos']]
        bloque(8, 'Comisiones — listas para importar',
               round(sum(r['monto_real'] for r in com_ok)), blue_fill, f'→ {len(com_ok)} factura(s)')
        if com_nd:
            bloque(9, 'Comisiones — sin datos (billing)',
                   round(sum(r['monto_real'] for r in com_nd)), red_fill, f'→ {len(com_nd)} factura(s)')

    # ── Cuentas sin datos ──────────────────────────────────────────────────────
    fila_det = 11
    if not df_nd_t.empty:
        c_lbl = ws_r.cell(row=fila_det, column=1, value='Cuentas sin datos:')
        c_lbl.font = Font(bold=True, name='Arial', size=10, color='CC0000')
        fila_det += 1
        seen_nd = set()
        for _, row in df_nd_t.iterrows():
            opid = row['operation_id']
            if opid in seen_nd:
                continue
            seen_nd.add(opid)
            op_rows_nd = df_nd_t[df_nd_t['operation_id'] == opid]
            nombre_nd = row['nombre_billing'] if row['nombre_billing'] else row['nombre_cuenta']
            monto_nd = calc_total_df(op_rows_nd)
            c1 = ws_r.cell(row=fila_det, column=1, value=f"• {nombre_nd} (ID {row['id_cuenta']})")
            c1.font = Font(name='Arial', size=10); c1.fill = red_fill
            c2 = ws_r.cell(row=fila_det, column=2, value=monto_nd)
            c2.font = Font(name='Arial', size=10); c2.fill = red_fill
            c2.number_format = '$#,##0'
            fila_det += 1
        fila_det += 1

    # ── Detalle por fecha ──────────────────────────────────────────────────────
    if not df_work.empty:
        c_lbl2 = ws_r.cell(row=fila_det, column=1, value='Detalle por fecha:')
        c_lbl2.font = Font(bold=True, name='Arial', size=11, color='1F4E79')
        fila_det += 1

        df_det = df_work.copy()
        df_det['_fecha_dt'] = df_det['fecha_compra'].apply(parse_fecha_dt)
        df_det = df_det.sort_values('_fecha_dt')

        for fecha_str, grupo_fecha in df_det.groupby('fecha_compra', sort=False):
            total_fecha = calc_total_df(grupo_fecha)
            # Encabezado de fecha
            c_fh = ws_r.cell(row=fila_det, column=1, value=f'📅 {fecha_str}')
            c_fh.font = Font(bold=True, name='Arial', size=11, color='FFFFFF')
            c_fh.fill = header_fill
            c_fh.alignment = Alignment(vertical='center')
            c_fv = ws_r.cell(row=fila_det, column=2, value=total_fecha)
            c_fv.font = Font(bold=True, name='Arial', size=11, color='FFFFFF')
            c_fv.fill = header_fill
            c_fv.number_format = '$#,##0'
            c_fv.alignment = Alignment(horizontal='center', vertical='center')
            ws_r.cell(row=fila_det, column=3).fill = header_fill
            ws_r.row_dimensions[fila_det].height = 20
            fila_det += 1

            # Filas individuales (una por operation_id único)
            seen_op = set()
            for _, row in grupo_fecha.iterrows():
                opid = row['operation_id']
                if opid in seen_op:
                    continue
                seen_op.add(opid)
                op_rows = grupo_fecha[grupo_fecha['operation_id'] == opid]
                nombre = row['nombre_billing'] if row['nombre_billing'] else row['nombre_cuenta']
                monto_op = calc_total_df(op_rows)
                row_fill = red_fill if row['sin_datos'] else PatternFill('solid', start_color='FFFFFF')
                c1 = ws_r.cell(row=fila_det, column=1, value=row['fecha_compra'])
                c1.font = Font(name='Arial', size=10); c1.fill = row_fill
                c2 = ws_r.cell(row=fila_det, column=2, value=f'Terminales - {nombre}')
                c2.font = Font(name='Arial', size=10); c2.fill = row_fill
                c3 = ws_r.cell(row=fila_det, column=3, value=monto_op)
                c3.font = Font(name='Arial', size=10); c3.fill = row_fill
                c3.number_format = '$#,##0'
                fila_det += 1
            fila_det += 1  # línea en blanco entre fechas

        fila_det += 1  # espacio antes de alertas

    if alertas_operador:
        ws_r.cell(row=fila_det, column=1,
            value=f'⚠ {len(alertas_operador)} pagos sin referencia:').font = Font(bold=True, name='Arial', size=10, color='CC5500')
        fila_det += 1
        for a in alertas_operador:
            for col, val in enumerate([a['fecha'],a['operador'],a['descripcion'],a['monto'],f"'{a['operation_id']}"], 1):
                cell = ws_r.cell(row=fila_det, column=col, value=val)
                cell.font = Font(name='Arial', size=9); cell.fill = PatternFill('solid', start_color='FFF2CC')
                if col == 4: cell.number_format = '$#,##0'
            fila_det += 1
        fila_det += 1

    if alertas_formato:
        ws_r.cell(row=fila_det, column=1,
            value=f'⚠ {len(alertas_formato)} filas con formato no reconocido:').font = Font(bold=True, name='Arial', size=10, color='CC5500')
        fila_det += 1
        for a in alertas_formato:
            for col, val in enumerate([a['fecha'],a['descripcion'],a['extref'],a['monto'],f"'{a['operation_id']}"], 1):
                cell = ws_r.cell(row=fila_det, column=col, value=val)
                cell.font = Font(name='Arial', size=9); cell.fill = PatternFill('solid', start_color='FFF2CC')
                if col == 4: cell.number_format = '$#,##0'
            fila_det += 1

    # ── Alertas Monto ──
    if alertas_monto:
        ws_al = wb.create_sheet('⚠ Alertas Monto')
        aplicar_header(ws_al,
            ['Cuenta','ID','Operation ID','Monto esperado','Monto real','Diferencia','Acción'],
            [30,10,18,16,16,14,45])
        for row_idx, a in enumerate(alertas_monto, 2):
            for col_idx, val in enumerate([a['nombre'],a['id_cuenta'],f"'{a['operation_id']}",
                                           a['monto_esperado'],a['monto_real'],a['diferencia'],
                                           'Verificar con ejecutivo'], 1):
                cell = ws_al.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name='Arial', size=10); cell.fill = yellow_fill
                if col_idx in [4,5,6]: cell.number_format = '$#,##0'

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output

# ─── Auditoría ────────────────────────────────────────────────────────────────
def run_auditoria(df_c, cols, billing_raw, refs, ids_facturados_real,
                  rl, slug_to_accid, slug_to_nombre):
    """
    Corre procesar() sin filtro de facturados para obtener el universo completo,
    luego cruza con el asiento real para clasificar cada operación.
    """
    # Universo completo (sin filtrar por ids_facturados)
    res = procesar(df_c, cols, billing_raw, refs, set(), rl,
                   slug_to_accid, slug_to_nombre,
                   hacer_terminales=True, hacer_comisiones=True)

    filas = []

    # ── Terminales ─────────────────────────────────────────────
    vistos = set()
    for row in res['rows']:
        opid = row['operation_id']
        if opid in vistos:
            continue
        vistos.add(opid)
        if opid in ids_facturados_real:
            estado = 'FACTURADO'
        elif row['sin_datos']:
            estado = 'SIN DATOS'
        else:
            estado = 'FALTANTE'
        nombre = row['nombre_billing'] or row['nombre_cuenta']
        filas.append({
            'operation_id': opid,
            'fecha':        row['fecha_compra'],
            'tipo':         'Terminal',
            'cuenta':       nombre,
            'id_cuenta':    row['id_cuenta'],
            'monto':        row['monto'],
            'estado':       estado,
            'detalle':      '' if estado != 'SIN DATOS' else 'Sin RUT en billing data',
        })

    # ── Deuda fija ─────────────────────────────────────────────
    for row in res['rows_comision']:
        opid = row['operation_id']
        if opid in ids_facturados_real:
            estado = 'FACTURADO'
        elif row['sin_datos']:
            estado = 'SIN DATOS'
        else:
            estado = 'FALTANTE'
        nombre = row['nombre_billing'] or row['nombre_cuenta']
        filas.append({
            'operation_id': opid,
            'fecha':        row['fecha_compra'],
            'tipo':         'Deuda fija',
            'cuenta':       nombre,
            'id_cuenta':    row['id_cuenta'],
            'monto':        row['monto_real'],
            'estado':       estado,
            'detalle':      '' if estado != 'SIN DATOS' else 'Sin RUT en billing data',
        })

    # ── Sin cuenta (deuda fija sin match en accounts.csv) ──────
    # Aun así puede estar facturada si el operation_id aparece en el asiento
    for row in res['comisiones_sin_cuenta']:
        opid_sc = str(row['operation_id']).replace("'", "").strip()
        if opid_sc in ids_facturados_real:
            estado_sc = 'FACTURADO'
            detalle_sc = f"Facturado (slug con typo en accounts.csv: {row['slug']})"
        else:
            estado_sc  = 'SIN CUENTA'
            detalle_sc = f"Slug no encontrado en accounts.csv: {row['slug']}"
        filas.append({
            'operation_id': row['operation_id'],
            'fecha':        row['fecha'],
            'tipo':         'Deuda fija',
            'cuenta':       row['extref'],
            'id_cuenta':    '',
            'monto':        row['monto'],
            'estado':       estado_sc,
            'detalle':      detalle_sc,
        })

    # ── Comerciales ────────────────────────────────────────────
    for row in res['alertas_operador']:
        filas.append({
            'operation_id': row['operation_id'],
            'fecha':        row['fecha'],
            'tipo':         'Terminal',
            'cuenta':       row['operador'],
            'id_cuenta':    '',
            'monto':        row['monto'],
            'estado':       'COMERCIAL',
            'detalle':      row['descripcion'],
        })

    # ── Formato desconocido ────────────────────────────────────
    for row in res['alertas_formato']:
        filas.append({
            'operation_id': row['operation_id'],
            'fecha':        row['fecha'],
            'tipo':         '?',
            'cuenta':       row['descripcion'][:60],
            'id_cuenta':    '',
            'monto':        row['monto'],
            'estado':       'FORMATO INVÁLIDO',
            'detalle':      row.get('extref', ''),
        })

    return filas


def generar_excel_auditoria(filas, periodo=''):
    wb  = Workbook()
    hoy = date.today().strftime('%d/%m/%Y')

    estado_fill = {
        'FACTURADO':       PatternFill('solid', start_color='C6EFCE'),
        'FALTANTE':        PatternFill('solid', start_color='FFD7D7'),
        'SIN DATOS':       PatternFill('solid', start_color='FFEB9C'),
        'SIN CUENTA':      PatternFill('solid', start_color='FFEB9C'),
        'COMERCIAL':       PatternFill('solid', start_color='DDEEFF'),
        'FORMATO INVÁLIDO':PatternFill('solid', start_color='F2F2F2'),
    }

    # ── Hoja principal: detalle completo ───────────────────────
    ws = wb.active
    ws.title = 'Auditoría completa'
    headers = ['Estado','Fecha','Tipo','Cuenta','ID Cuenta','Operation ID','Monto','Detalle']
    widths  = [18, 14, 12, 42, 14, 22, 14, 50]
    aplicar_header(ws, headers, widths)

    for r, f in enumerate(filas, 2):
        vals = [f['estado'], f['fecha'], f['tipo'], f['cuenta'],
                f['id_cuenta'], f"'{f['operation_id']}", f['monto'], f['detalle']]
        fill = estado_fill.get(f['estado'], PatternFill())
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            if c == 7: cell.number_format = '$#,##0'

    # ── Hoja faltantes ─────────────────────────────────────────
    faltantes = [f for f in filas if f['estado'] == 'FALTANTE']
    if faltantes:
        ws_f = wb.create_sheet('⚠ Faltantes')
        aplicar_header(ws_f, ['Fecha','Tipo','Cuenta','ID Cuenta','Operation ID','Monto'], [14,12,42,14,22,14])
        for r, f in enumerate(faltantes, 2):
            vals = [f['fecha'], f['tipo'], f['cuenta'], f['id_cuenta'],
                    f"'{f['operation_id']}", f['monto']]
            for c, v in enumerate(vals, 1):
                cell = ws_f.cell(row=r, column=c, value=v)
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill('solid', start_color='FFD7D7')
                if c == 6: cell.number_format = '$#,##0'

    # ── Hoja resumen ───────────────────────────────────────────
    ws_r = wb.create_sheet('Resumen auditoría')
    ws_r.column_dimensions['A'].width = 35
    ws_r.column_dimensions['B'].width = 16
    ws_r.column_dimensions['C'].width = 20

    t = ws_r.cell(row=1, column=1, value='AUDITORÍA DE FACTURACIÓN')
    t.font = Font(bold=True, name='Arial', size=13, color='FFFFFF')
    t.fill = header_fill; t.alignment = Alignment(horizontal='center', vertical='center')
    ws_r.merge_cells('A1:C1'); ws_r.row_dimensions[1].height = 24

    ws_r.cell(row=2, column=1,
        value=f"Generado: {hoy}" + (f"  |  Período: {periodo}" if periodo else '')
    ).font = Font(italic=True, name='Arial', size=10, color='666666')


    conteo = Counter(f['estado'] for f in filas)
    total  = len({f['operation_id'] for f in filas})

    def fila_r(r, label, valor, color, nota=''):
        c1 = ws_r.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, name='Arial', size=11)
        c1.fill = PatternFill('solid', start_color=color)
        c1.alignment = Alignment(horizontal='right', vertical='center')
        c1.border = Border(outline=Side(style='thin'))
        c2 = ws_r.cell(row=r, column=2, value=valor)
        c2.font = Font(bold=True, name='Arial', size=12, color='1F4E79')
        c2.fill = PatternFill('solid', start_color=color)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border = Border(outline=Side(style='thin'))
        if nota:
            ws_r.cell(row=r, column=3, value=nota).font = Font(italic=True, name='Arial', size=9, color='444444')

    fila_r(4,  'Total operaciones verificadas', total,                        'D9E1F2')
    fila_r(6,  '✓  Correctamente facturadas',   conteo.get('FACTURADO', 0),   'C6EFCE',
           f'{round(conteo.get("FACTURADO",0)/total*100) if total else 0}% del total')
    fila_r(7,  '✗  Faltantes por facturar',      conteo.get('FALTANTE', 0),    'FFD7D7',
           '← Revisar urgente' if conteo.get('FALTANTE', 0) else '')
    fila_r(8,  '⚠  Sin datos de facturación',    conteo.get('SIN DATOS', 0) + conteo.get('SIN CUENTA', 0), 'FFEB9C',
           'Pedir RUT/cuenta al ejecutivo' if conteo.get('SIN DATOS',0)+conteo.get('SIN CUENTA',0) else '')
    fila_r(9,  'ℹ  Comerciales (no facturable)', conteo.get('COMERCIAL', 0),  'DDEEFF')
    fila_r(10, '?  Formato no reconocido',        conteo.get('FORMATO INVÁLIDO', 0), 'F2F2F2')

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output


# ─── UI Principal ─────────────────────────────────────────────────────────────
def main():
    if not check_access():
        return

    rl = cargar_regiones()

    # ── Header ─────────────────────────────────────────────────
    st.markdown("""
    <div style="background:#FF5023; border-radius:10px; padding:22px 28px 16px; margin-bottom:8px;">
        <span style="font-family:'Barlow',sans-serif; font-size:26px; font-weight:700; color:#FFFFFF; letter-spacing:-0.5px;">
            🧾 Facturación Terminales
        </span><br>
        <span style="font-family:'Barlow',sans-serif; font-size:13px; color:#FFC9BB; font-weight:400;">
            Anser Indicus SPA — Procesador automático de cobranzas Mercado Pago
        </span>
    </div>
    """, unsafe_allow_html=True)

    # ── Guía de uso ────────────────────────────────────────────
    with st.expander("📖 Guía de uso — cómo funciona el procesador"):
        st.markdown("""
        ### Flujo de trabajo

        **PASO 1 — Crear contactos**
        Usalo cuando hay cuentas nuevas que todavía no están en Odoo.
        1. Subí los 4 archivos requeridos y hacé clic en **Procesar**
        2. Descargá el archivo `contactos_FECHA.xlsx` que genera el procesador
        3. Importalo en Odoo: **Contactos → ⚙️ → Importar registros**
        4. Una vez importado, **exportá el res.partner de nuevo desde Odoo** (ahora tiene los DB_IDs nuevos)
        5. Con ese archivo actualizado, pasá al PASO 2

        **PASO 2 — Facturar**
        Usalo cuando todos los contactos ya están en Odoo con su DB_ID.
        1. Seleccioná qué querés facturar: **Terminales** o **Deuda fija**
        2. Subí los 4 archivos (con el res.partner actualizado de Odoo)
        3. Si facturás **Deuda fija**, subí también el Accounts CSV
        4. Hacé clic en **Procesar** y descargá `facturar_terminales_FECHA.xlsx`
        5. Importalo en Odoo para crear las facturas

        ---
        ### Archivos requeridos

        | Archivo | Dónde lo obtenés |
        |---|---|
        | **Collection Mercado Pago** | Portal Mercado Pago → Actividades → Exportar |
        | **Billing data** | dash.fu.do → Exportar billing |
        | **Contactos Odoo** (res.partner) | Odoo → Contactos → Exportar → res.partner |
        | **Asiento contable** | Odoo → Contabilidad → Asientos → Exportar |
        | **Accounts CSV** *(solo Deuda fija)* | dash.fu.do → Cuentas → Exportar |

        ---
        ### Colores en los archivos Excel

        | Color | Significado |
        |---|---|
        | 🔴 Fila roja | Cuenta sin datos de facturación (sin RUT en billing) |
        | 🟠 Celda naranja | Sin DB_ID en Odoo — no se puede importar todavía |
        | 🟡 Celda amarilla | Verificar manualmente (monto diferente u otro aviso) |
        | 🟢 OK / verde | Todo en orden |
        | 🔴 SUPERA LÍMITE SII | La comuna tiene más de 20 caracteres — SII rechazará la factura |
        """)

    st.divider()

    # ── Sidebar ────────────────────────────────────────────────
    with st.sidebar:
        st.header("¿Qué vas a hacer?")

        paso = st.radio("", [
            "📋  PASO 1 — Crear contactos",
            "🧾  PASO 2 — Facturar",
            "🔍  PASO 3 — Auditoría",
        ], label_visibility="collapsed")

        es_paso1 = "PASO 1" in paso

        tipo_fact = "Terminales + Deuda fija"  # default (para PASO 1 procesar todo)
        if not es_paso1:
            st.markdown("**¿Qué querés facturar?**")
            tipo_fact = st.selectbox("", [
                "Terminales",
                "Deuda fija",
                "Terminales + Deuda fija",
            ], label_visibility="collapsed")

        st.divider()
        st.markdown("""
        **Leyenda Excel:**
        🔴 sin datos · 🟠 sin DB_ID
        🟡 verificar · 🟢 OK
        """)
        user_email = st.session_state.get("user_email", "")
        if user_email:
            st.markdown(
                f'<p style="font-family:Barlow,sans-serif; font-size:12px; '
                f'color:#FFC9BB; margin:0 0 2px 0;">👤 {user_email}</p>',
                unsafe_allow_html=True
            )
        st.markdown(
            '<p style="font-family:Barlow,sans-serif; font-size:11px; '
            'color:#9B9BD4; margin:0;">v1.0 · Facturación Terminales</p>',
            unsafe_allow_html=True
        )

    hacer_terminales = "Terminales" in tipo_fact
    hacer_comisiones = "Deuda fija" in tipo_fact
    proc_term = hacer_terminales or es_paso1
    proc_com  = hacer_comisiones or es_paso1

    # ── File uploaders ──────────────────────────────────────────
    st.subheader("📁 Archivos requeridos")
    col1, col2 = st.columns(2)
    with col1:
        f_col = st.file_uploader(
            "Collection Mercado Pago  *(requerido)*", type=['xlsx','csv'],
            help="collection-FECHA.xlsx exportado de Mercado Pago")
        f_bil = st.file_uploader(
            "Billing data  *(requerido)*", type=['csv','xlsx'],
            help="billing_data_FECHA.csv exportado de dash.fu.do")
    with col2:
        f_con = st.file_uploader(
            "Contactos Odoo — res.partner  *(requerido)*", type=['xlsx'],
            help="Exportación del modelo res.partner desde Odoo")
        f_odo = st.file_uploader(
            "Asiento contable Odoo  *(requerido)*", type=['xlsx'],
            help="Para detectar facturas ya procesadas y evitar duplicados")

    f_acc = None
    if proc_com:
        st.subheader("📁 Adicional para Deuda fija")
        f_acc = st.file_uploader(
            "Accounts CSV  *(necesario para Deuda fija)*", type=['csv'],
            help="accounts_FECHA.csv exportado de dash.fu.do")

    # ── Botón ───────────────────────────────────────────────────
    st.divider()
    archivos_ok = all([f_col, f_bil, f_con, f_odo])
    if not archivos_ok:
        faltantes = [n for f, n in [(f_col,"Collection"),(f_bil,"Billing data"),
                                    (f_con,"Contactos Odoo"),(f_odo,"Asiento contable")] if not f]
        st.info(f"⬆️ Subí los archivos requeridos: **{', '.join(faltantes)}**")

    boton = st.button("🚀  Procesar", type="primary", use_container_width=True, disabled=not archivos_ok)

    if boton:
        # ── Lectura ──────────────────────────────────────────────
        with st.spinner("Leyendo archivos..."):
            try:
                df_c, cols = leer_collection(f_col)
            except Exception as e:
                st.error(f"❌ **Collection** ({f_col.name}): {e}"); return
            try:
                billing_raw = leer_billing(f_bil)
            except Exception as e:
                st.error(f"❌ **Billing data** ({f_bil.name}): {e}"); return
            try:
                refs = leer_contactos(f_con)
            except Exception as e:
                st.error(f"❌ **Contactos Odoo** ({f_con.name}): {e}"); return
            try:
                ids_facturados = leer_odoo(f_odo)
            except Exception as e:
                st.error(f"❌ **Asiento contable** ({f_odo.name}): {e}"); return
            try:
                slug_to_accid, slug_to_nombre = leer_accounts(f_acc) if f_acc else ({}, {})
            except Exception as e:
                st.warning(f"⚠️ **Accounts CSV**: {e} — se continúa sin Deuda fija")
                slug_to_accid, slug_to_nombre = {}, {}

        # ── Procesamiento ────────────────────────────────────────
        with st.spinner("Procesando collection..."):
            resultado = procesar(
                df_c, cols, billing_raw, refs, ids_facturados, rl,
                slug_to_accid, slug_to_nombre,
                hacer_terminales=proc_term,
                hacer_comisiones=proc_com,
            )
            st.session_state['resultado'] = resultado

        rows          = resultado['rows']
        rows_comision = resultado['rows_comision']
        dup_count     = len(resultado['duplicados'])
        alertas_monto = resultado['alertas_monto']
        alertas_op    = resultado['alertas_operador']
        alertas_fmt   = resultado['alertas_formato']
        csc           = resultado['comisiones_sin_cuenta']

        df_work     = pd.DataFrame(rows)
        df_comision = pd.DataFrame(rows_comision)

        if df_work.empty and df_comision.empty and not csc:
            st.warning("⚠️ No se encontraron filas para procesar en el collection.")
            return

        with st.spinner("Clasificando contactos..."):
            casos_ok, casos_dc, casos_act, casos_crear, casos_rut_otro = clasificar_contactos(
                df_work, df_comision, refs, rl
            )

        hay_contactos_nuevos   = len(casos_crear) > 0
        hay_acciones_contactos = hay_contactos_nuevos or len(casos_act) > 0 or len(casos_dc) > 0

        # ── Métricas ──────────────────────────────────────────────
        st.divider()
        st.subheader("📊 Resultado")
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        total_term = calc_total_df(df_work)
        total_com  = round(sum(r['monto_real'] for r in rows_comision)) if rows_comision else 0
        col_m1.metric("Terminales", f"${total_term:,.0f}",
                      f"{df_work['operation_id'].nunique() if not df_work.empty else 0} facturas")
        col_m2.metric("Deuda fija", f"${total_com:,.0f}", f"{len(rows_comision)} facturas")
        col_m3.metric("TOTAL", f"${total_term + total_com:,.0f}")
        col_m4.metric("Duplicados excluidos", dup_count)

        # Alertas generales
        if casos_act:
            st.warning(f"⚠️ **{len(casos_act)} contacto(s)** con RUT cambiado — ver hoja '⚠ Contactos a actualizar'.")
        if casos_dc:
            st.info(f"ℹ️ **{len(casos_dc)} contacto(s)** con datos actualizados en billing — ver hoja '⚠ Datos actualizados'.")
        if alertas_op:
            st.warning(f"⚠️ **{len(alertas_op)} pago(s)** de comerciales sin referencia — ver hoja Resumen.")
        if alertas_fmt:
            st.warning(f"⚠️ **{len(alertas_fmt)} fila(s)** con formato desconocido — ver hoja Resumen.")
        if csc:
            st.warning(f"⚠️ **{len(csc)} Deuda(s) fija(s)** sin cuenta resuelta — ver hoja '⚠ Comisiones sin cuenta'.")
        if alertas_monto:
            st.warning(f"⚠️ **{len(alertas_monto)} alerta(s)** de monto — ver hoja '⚠ Alertas Monto'.")

        # ── Descargas ─────────────────────────────────────────────
        st.divider()
        st.subheader("📥 Descargar archivos")
        dl_col1, dl_col2 = st.columns(2)

        # Archivo de contactos (siempre que haya algo que reportar)
        if hay_acciones_contactos:
            with st.spinner("Generando contactos.xlsx..."):
                excel_cont = generar_excel_contactos(casos_crear, casos_act, casos_rut_otro, casos_dc, rl)
            nombre_cont = f"contactos_{date.today().strftime('%Y%m%d')}.xlsx"
            with dl_col1:
                if hay_contactos_nuevos:
                    st.warning(f"⚠️ {len(casos_crear)} contacto(s) nuevo(s) — importalos en Odoo")
                st.download_button(
                    label=f"📋 {nombre_cont}",
                    data=excel_cont,
                    file_name=nombre_cont,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Importar en Odoo antes de facturar"
                )
        else:
            with dl_col1:
                st.success("✅ Todos los contactos ya están en Odoo")

        # Archivo de facturación / instrucciones
        if es_paso1:
            with dl_col2:
                st.info(
                    "**PASO 1 completado.**\n\n"
                    "1. Descargá el archivo de contactos ←\n"
                    "2. Importalo en Odoo (Contactos → Importar)\n"
                    "3. Exportá el res.partner actualizado desde Odoo\n"
                    "4. Volvé aquí, seleccioná **PASO 2** y subí el nuevo res.partner"
                )
        elif hay_contactos_nuevos:
            with dl_col2:
                # Mensaje claro explicando por qué no se generó la facturación
                faltantes_lista = "\n".join(
                    f"- **{c['nombre_cuenta']}** (ID: {c['id_cuenta']})"
                    for c in casos_crear[:10]
                )
                if len(casos_crear) > 10:
                    faltantes_lista += f"\n- ... y {len(casos_crear)-10} más"
                st.error(
                    f"⛔ **No se generó el archivo de facturación.**\n\n"
                    f"El archivo de contactos de Odoo que subiste "
                    f"(**{f_con.name}**) no incluye {len(casos_crear)} cuenta(s) "
                    f"que aparecen en el collection. "
                    f"Esto ocurre cuando subiste un res.partner anterior a la importación "
                    f"de los contactos nuevos.\n\n"
                    f"**Cuentas que faltan en Odoo:**\n{faltantes_lista}"
                )
                st.info(
                    "**Qué tenés que hacer:**\n"
                    "1. Descargá el archivo de contactos ← (botón de la izquierda)\n"
                    "2. Importalo en Odoo (Contactos → Importar registros)\n"
                    "3. En Odoo, exportá el res.partner de nuevo\n"
                    "4. Subí ese nuevo export en el campo **Contactos Odoo** (arriba)\n"
                    "5. Hacé clic en **Procesar** de nuevo"
                )
        elif df_work.empty and not rows_comision:
            with dl_col2:
                st.info("No hay filas para facturar en este collection.")
        else:
            empty_cols = ['id_cuenta','cantidad','descuento','nombre_cuenta','operation_id',
                          'monto','RUT_billing','RUT_odoo','razon_social','nombre_billing',
                          'giro','domicilio','comuna','email','db_id','contacto_nombre',
                          'monto_diferente','sin_datos','es_consumidor_final','fecha_compra']
            with st.spinner("Generando Excel de facturación..."):
                excel_fact = generar_excel_facturacion(
                    df_work if not df_work.empty else pd.DataFrame(columns=empty_cols),
                    rows_comision if hacer_comisiones else [],
                    alertas_monto, alertas_op, alertas_fmt,
                    comisiones_sin_cuenta=csc,
                )
            nombre_fact = f"facturar_terminales_{date.today().strftime('%Y%m%d')}.xlsx"
            with dl_col2:
                st.download_button(
                    label=f"🧾 {nombre_fact}",
                    data=excel_fact,
                    file_name=nombre_fact,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    help="Importar en Odoo para crear las facturas"
                )

        st.success("✅ Procesamiento completado")

    # ══════════════════════════════════════════════════════════
    # PASO 3 — Auditoría de integridad
    # ══════════════════════════════════════════════════════════
    if "PASO 3" in paso:
        st.markdown("""
        <div style="background:#3938A0; border-radius:10px; padding:18px 24px 14px; margin-bottom:12px;">
            <div style="font-family:'Barlow',sans-serif; font-size:20px; font-weight:700; color:#FFFFFF;">
                🔍 Auditoría de facturación
            </div>
            <div style="font-family:'Barlow',sans-serif; font-size:13px; color:#E1E1F5; margin-top:4px;">
                Verificá que todas las operaciones del período estén correctamente facturadas
            </div>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("ℹ️ ¿Cómo funciona la auditoría?"):
            st.markdown("""
            La herramienta analiza el collection del período y lo cruza con el asiento contable de Odoo para clasificar cada operación:

            | Estado | Significado |
            |---|---|
            | ✓ **FACTURADO** | El pago aparece en el asiento contable — OK |
            | ✗ **FALTANTE** | El pago debería facturarse pero no está en Odoo |
            | ⚠ **SIN DATOS** | Sin RUT en billing data — no se puede facturar hasta que el ejecutivo lo complete |
            | ⚠ **SIN CUENTA** | Deuda fija sin match en accounts.csv |
            | ℹ **COMERCIAL** | Pago de comercial — no se factura |
            | ? **FORMATO INVÁLIDO** | No se pudo interpretar la descripción del pago |

            **Qué necesitás:** el collection del período completo + billing data + asiento contable de Odoo.
            El asiento contable debería incluir **todos los asientos del período** para que la comparación sea exacta.
            """)

        st.subheader("📁 Archivos para la auditoría")
        periodo_txt = st.text_input("Período a auditar (opcional)", placeholder="Ej: Julio 2026")

        col_a1, col_a2 = st.columns(2)
        with col_a1:
            fa_col = st.file_uploader("Collection Mercado Pago *(requerido)*",
                                      type=['xlsx','csv'], key="aud_col")
            fa_bil = st.file_uploader("Billing data *(requerido)*",
                                      type=['csv','xlsx'], key="aud_bil")
        with col_a2:
            fa_con = st.file_uploader("Contactos Odoo — res.partner *(requerido)*",
                                      type=['xlsx'], key="aud_con")
            fa_odo = st.file_uploader("Asiento contable Odoo *(requerido)*",
                                      type=['xlsx'], key="aud_odo")

        fa_acc = st.file_uploader(
            "Accounts CSV *(opcional — para auditar Deuda fija)*",
            type=['csv'], key="aud_acc")

        st.divider()
        aud_ok = all([fa_col, fa_bil, fa_con, fa_odo])
        if not aud_ok:
            falt_a = [n for f, n in [(fa_col,"Collection"),(fa_bil,"Billing data"),
                                     (fa_con,"Contactos Odoo"),(fa_odo,"Asiento contable")] if not f]
            st.info(f"⬆️ Subí los archivos requeridos: **{', '.join(falt_a)}**")

        if st.button("🔍  Analizar", type="primary", use_container_width=True, disabled=not aud_ok):
            with st.spinner("Leyendo archivos..."):
                try:    df_ac, acols = leer_collection(fa_col)
                except Exception as e: st.error(f"❌ Collection: {e}"); st.stop()
                try:    abil = leer_billing(fa_bil)
                except Exception as e: st.error(f"❌ Billing data: {e}"); st.stop()
                try:    arefs = leer_contactos(fa_con)
                except Exception as e: st.error(f"❌ Contactos Odoo: {e}"); st.stop()
                try:    aids_real = leer_odoo(fa_odo)
                except Exception as e: st.error(f"❌ Asiento contable: {e}"); st.stop()
                if len(aids_real) < 5:
                    st.warning(f"⚠️ El asiento contable tiene solo **{len(aids_real)} entrada(s)**. "
                               f"Para una auditoría completa necesitás exportar el asiento con "
                               f"**todas las facturas del período**, no solo una.")
                try:    aslug_id, aslug_nom = leer_accounts(fa_acc) if fa_acc else ({}, {})
                except Exception as e:
                    st.warning(f"⚠️ Accounts CSV: {e} — auditoría de Deuda fija incompleta")
                    aslug_id, aslug_nom = {}, {}

            with st.spinner("Analizando operaciones..."):
                filas_aud = run_auditoria(df_ac, acols, abil, arefs, aids_real,
                                          rl, aslug_id, aslug_nom)

            if not filas_aud:
                st.warning("No se encontraron operaciones para auditar en este collection.")
                st.stop()

            # ── Métricas resumen ────────────────────────────────────
        
            conteo_a = Counter(f['estado'] for f in filas_aud)
            total_a  = len({f['operation_id'] for f in filas_aud})
            n_fact   = conteo_a.get('FACTURADO', 0)
            n_falt   = conteo_a.get('FALTANTE', 0)
            n_sdat   = conteo_a.get('SIN DATOS', 0) + conteo_a.get('SIN CUENTA', 0)
            n_otros  = conteo_a.get('COMERCIAL', 0) + conteo_a.get('FORMATO INVÁLIDO', 0)

            st.divider()
            st.subheader("📊 Resultado de la auditoría")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total verificadas", total_a)
            m2.metric("✓ Facturadas", n_fact,
                      f"{round(n_fact/total_a*100) if total_a else 0}%")
            m3.metric("✗ Faltantes", n_falt,
                      "← revisar" if n_falt else "ninguna")
            m4.metric("⚠ Inconsistencias", n_sdat)

            # ── Alertas ─────────────────────────────────────────────
            if n_falt == 0 and n_sdat == 0:
                msg_ok = "✅ Todo el período está correctamente facturado. No hay operaciones pendientes."
                if n_otros:
                    msg_ok += f" Además, hay {n_otros} pago(s) que no corresponden a terminales ni deuda fija — ver detalle abajo."
                st.success(msg_ok)
            else:
                if n_falt:
                    falt_list = [f for f in filas_aud if f['estado'] == 'FALTANTE']
                    with st.expander(f"✗ **{n_falt} operación(es) FALTANTE(S) — sin facturar**", expanded=True):
                        for f in falt_list:
                            st.markdown(
                                f"- **{f['cuenta']}** (ID: {f['id_cuenta']}) — "
                                f"{f['tipo']} — {f['fecha']} — "
                                f"**${f['monto']:,.0f}**  `{f['operation_id']}`"
                            )
                if n_sdat:
                    sdat_list = [f for f in filas_aud if f['estado'] in ('SIN DATOS','SIN CUENTA')]
                    with st.expander(f"⚠ **{n_sdat} operación(es) sin datos — no se pueden facturar**"):
                        for f in sdat_list:
                            st.markdown(
                                f"- **{f['cuenta']}** — {f['tipo']} — {f['fecha']} — "
                                f"${f['monto']:,.0f} — _{f['detalle']}_"
                            )

            # Comerciales y formato inválido — siempre mostrar si hay
            if n_otros:
                otros_list = [f for f in filas_aud if f['estado'] in ('COMERCIAL', 'FORMATO INVÁLIDO')]
                with st.expander(f"ℹ️ **{n_otros} pago(s) excluido(s) de la facturación**"):
                    for f in otros_list:
                        etiqueta = "Comercial" if f['estado'] == 'COMERCIAL' else "Formato inválido"
                        st.markdown(
                            f"- **{etiqueta}** — {f['cuenta']} — {f['fecha']} — "
                            f"${float(f['monto']):,.0f}  `{f['operation_id']}`"
                            + (f" — _{f['detalle']}_" if f['detalle'] else "")
                        )

            # ── Descarga ────────────────────────────────────────────
            st.divider()
            with st.spinner("Generando Excel de auditoría..."):
                excel_aud = generar_excel_auditoria(filas_aud, periodo=periodo_txt)
            nombre_aud = f"auditoria_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label=f"📥 Descargar {nombre_aud}",
                data=excel_aud,
                file_name=nombre_aud,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

if __name__ == '__main__':
    main()
